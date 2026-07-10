"""캐나다 LMIA 승인 고용주 명부 대조 검증 도구(오프라인, 일회성/주기적).

레지스트리 회사를 ESDC 'Positive LMIA Employers List' CSV 와 정밀 매칭해 후보를
출력한다. 사람이 검토 후 companies.json 에 "ca_sponsor": true 를 수동 반영한다.
런타임 ETL 과 무관. uk/h1b/ind 검증 스크립트와 동일 구조.

명부 출처(분기별 파일 ~1.2MB, open.canada.ca):
  https://open.canada.ca/data/en/dataset/90fed587-1364-4f33-a9ee-208181dc0b97
  최신 파일은 XLSX(.xlsx) — 이 스크립트가 직접 읽는다(구형 .xls 는 CSV 로 변환 필요).
  (개인명 고용주는 제외되어 완전하지 않음 — false negative 는 있어도 오검증은 없음)

사용:
  python scripts/verify_ca_sponsors.py /path/to/tfwp_2025q3_pos_en.xlsx  # 또는 .csv
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

# 정규화·매칭은 공유 모듈(이름 + 위치 disambiguation + confidence)로 통일. uk/h1b/ind 와 동일.
from sponsor_match import company_names, find_candidates

REGISTRY_PATH = Path(__file__).parent.parent / "dev_jobs_core" / "data" / "companies.json"


def _rows(path: str):
    """CSV 또는 XLSX 를 dict 행(헤더→값)으로 yield. 확장자로 판별. XLSX 는 openpyxl(이미 의존성)."""
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows, [])]
        for r in rows:
            yield {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        wb.close()
        return
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        yield from csv.DictReader(f)


def _row_location(row: dict) -> str:
    """LMIA 명부 행에서 위치(주/준주 + 주소/도시) 추출. 헤더 변형에 관대하게."""
    province = extra = ""
    for k, v in row.items():
        kl = (k or "").lower()
        val = str(v).strip() if v is not None else ""
        if "province" in kl or "territory" in kl:
            province = val
        elif "address" in kl or "city" in kl or "location" in kl:
            extra = val
    return " ".join(p for p in (extra, province) if p)


def _load_register(path: str) -> list[tuple[str, str]]:
    """명부 → (고용주명, 위치) 목록. 헤더 변형(Employer/Business Operating Name 등)에 관대하게."""
    out = []
    for row in _rows(path):
        org = ""
        for k, v in row.items():
            kl = (k or "").lower()
            if "employer" in kl or "operating name" in kl or "business name" in kl:
                org = str(v).strip() if v is not None else ""
                if org:
                    break
        if org:
            out.append((org, _row_location(row)))
    return out


def main():
    if len(sys.argv) < 2:
        raise SystemExit("사용: python scripts/verify_ca_sponsors.py /path/to/lmia_employers.csv")
    register = _load_register(sys.argv[1])
    with open(REGISTRY_PATH, encoding="utf-8") as f:
        data = json.load(f)
    cos = {k: v for k, v in data.items() if not k.startswith("_")}

    hits = []
    for name, info in cos.items():
        cands = find_candidates(company_names(name, info), info.get("hq"), register)
        if cands:
            hits.append((name, info, cands[0]))

    print(f"=== CA LMIA 명부 매칭 후보: {len(hits)}/{len(cos)} (검토 후 companies.json 반영) ===")
    print("    confidence: high=이름+위치일치 / medium=단독 이름일치 / low=동명 모호(검토 필수)")
    for name, info, c in sorted(hits, key=lambda h: ({"high": 0, "medium": 1, "low": 2}[h[2].confidence], h[0])):
        already = " [이미 플래그됨]" if info.get("ca_sponsor") is True else ""
        dom = f" <{info['domain']}>" if info.get("domain") else ""
        loc = f" @{c.loc}" if c.loc else ""
        print(f"  [{c.confidence:<6}] {name:<16}{dom} [{info.get('ats')}] -> {c.org}{loc}{already}")


if __name__ == "__main__":
    main()
